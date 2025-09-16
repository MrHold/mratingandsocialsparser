from fastapi import FastAPI, APIRouter
from fastapi.responses import FileResponse
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment
import tempfile
from contextlib import asynccontextmanager

from .services.data_collect import collect_uni_stats
from .services.parsers.mratingparser import get_unis

from .models.schemas import AnalyzeRequest

from pyrogram import Client
import os
from dotenv import load_dotenv

load_dotenv()

@asynccontextmanager
async def lifespan(app: FastAPI):
    await startup()
    yield
    await shutdown()

app = FastAPI(lifespan=lifespan)
router = APIRouter()

tg_api_id = os.environ.get("TG_API_ID")
tg_api_hash = os.environ.get("TG_API_HASH")
tg_session_name = os.environ.get("TG_SESSION_NAME")
tg_session_string = os.environ.get("TG_SESSION_STRING")

async def startup():
    global app_client
    app_client = Client(
        name="my_userbot",
        api_id=tg_api_id,
        api_hash=tg_api_hash,
        session_string=tg_session_string,
        in_memory=True
    )
    await app_client.start()

async def shutdown():
    global app_client
    await app_client.stop()

@app.get("/")
def read_root():
    return {"Hello": "World"}


@router.post("/analyze")
async def analyze_endpoint(data: AnalyzeRequest):
    universities = [x.webname for x in data.universities]
    results = {}
    for webname in universities:
        results[webname] = await collect_uni_stats(
            webname, data.start_date, data.end_date, data.socials, app_client
        )

    return results


@router.post("/export")
async def export_endpoint(data: AnalyzeRequest):
    res = await analyze_endpoint(data)
    print(res)
    wb = Workbook()
    ws = wb.active
    headers = [
        "Месяц",
        "М-рейтинг",
        "М-рейтинг СоцСети",
        "М-рейтинг ВК",
        "М-рейтинг ТГ",
        "М-рейтинг Рутуб",
        "М-рейтинг Сайт",
        "М-рейтинг СМИ",
        "Публикаций ВК",
        "Просмотров ВК",
        "Реакций ВК",
        "Репостов ВК",
        "Комментариев ВК",
        "Публикаций ТГ",
        "Просмотров ТГ",
        "Реакций ТГ",
        "Репостов ТГ",
        "Комментариев ТГ",
        "Публикаций Рутуб",
        "Просмотров Рутуб",
    ]
    for i in res:
        ws.append([res[i]["university"]])
        ws.merge_cells(
            start_row=ws.max_row,
            start_column=1,
            end_row=ws.max_row,
            end_column=len(headers),
        )
        ws.cell(row=ws.max_row, column=1).alignment = Alignment(horizontal="center")
        ws.append(headers)
        for month_data in res[i]["data"]:
            row = [
                month_data["month"],
                month_data["mrating"],
                month_data["mrating_socials"],
                month_data["mrating_vk"],
                month_data["mrating_tg"],
                month_data["mrating_rutube"],
                month_data["mrating_site"],
                month_data["mrating_smi"],
                month_data["posts_vk"],
                month_data["views_vk"],
                month_data["reacts_vk"],
                month_data["reposts_vk"],
                month_data["comments_vk"],
                month_data["posts_tg"],
                month_data["views_tg"],
                month_data["reacts_tg"],
                month_data["forwards_tg"],
                month_data["comments_tg"],
                month_data["posts_rutube"],
                month_data["views_rutube"],
            ]
            ws.append(row)
        ws.append([])
    # Увеличение ширины столбцов, чтобы помещались заголовки и значения
    for header in range(len(headers)):
        ws.column_dimensions[chr(65 + header)].width = len(headers[header]) + 2

    with tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp:
        wb.save(tmp.name)
        temp_path = tmp.name

    return FileResponse(
        path=temp_path,
        filename=f"статистика{datetime.now().strftime('%Y-%m-%d-%H-%M-%S')}.xlsx",
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/universities")
async def universities():
    return get_unis()


app.include_router(router)
